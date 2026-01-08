"""
Excel Workbook (.xlsx) operations using openpyxl
"""

from pathlib import Path
from typing import Optional, List, Any

from openpyxl import load_workbook, Workbook


class ExcelWorkbook:
    """Operations for Excel workbooks"""

    def __init__(self, path: str):
        self.path = Path(path).expanduser().resolve()
        self._wb = None

    def _load(self, read_only: bool = False, data_only: bool = False):
        if not self._wb:
            self._wb = load_workbook(str(self.path), read_only=read_only, data_only=data_only)
        return self._wb

    def _save(self, output: Optional[str] = None):
        out_path = Path(output).expanduser().resolve() if output else self.path
        self._wb.save(str(out_path))
        return out_path

    def close(self):
        if self._wb:
            self._wb.close()
            self._wb = None

    @staticmethod
    def create(path: str, sheet_name: str = "Sheet1", headers: Optional[List[str]] = None) -> 'ExcelWorkbook':
        """Create a new Excel workbook"""
        file_path = Path(path).expanduser().resolve()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if headers:
            ws.append(headers)

        wb.save(str(file_path))
        wb.close()
        return ExcelWorkbook(str(file_path))

    def get_sheets(self) -> List[str]:
        """Get list of sheet names"""
        wb = self._load(read_only=True)
        return wb.sheetnames

    def read_sheet(self, sheet: Optional[str] = None, cell_range: Optional[str] = None) -> List[List[Any]]:
        """Read data from a sheet"""
        wb = self._load(read_only=True, data_only=True)

        if sheet:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet]
        else:
            ws = wb.active

        if cell_range:
            data = [[cell.value for cell in row] for row in ws[cell_range]]
        else:
            data = [[cell.value for cell in row] for row in ws.iter_rows()]

        return data

    def to_markdown(self, sheet: Optional[str] = None) -> str:
        """Convert sheet to markdown table"""
        data = self.read_sheet(sheet)
        wb = self._load(read_only=True)
        ws = wb[sheet] if sheet else wb.active

        result = f"**Sheet:** {ws.title}\n\n"

        if data and data[0]:
            headers = [str(h) if h else "" for h in data[0]]
            result += "| " + " | ".join(headers) + " |\n"
            result += "| " + " | ".join(["---"] * len(headers)) + " |\n"

            for row in data[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                while len(cells) < len(headers):
                    cells.append("")
                result += "| " + " | ".join(cells[:len(headers)]) + " |\n"

        return result

    def write_cell(self, sheet: str, cell: str, value: Any) -> Any:
        """Write a value to a cell, returns old value"""
        wb = self._load()

        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")

        ws = wb[sheet]
        old_value = ws[cell].value
        ws[cell] = value

        return old_value

    def write_row(self, sheet: str, row: int, values: List[Any], start_col: str = "A"):
        """Write values to a row"""
        wb = self._load()

        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found.")

        ws = wb[sheet]
        start_col_idx = ord(start_col.upper()) - ord('A') + 1

        for i, val in enumerate(values):
            ws.cell(row=row, column=start_col_idx + i, value=val)

    def append_row(self, sheet: str, values: List[Any]) -> int:
        """Append a row, returns new row number"""
        wb = self._load()

        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found.")

        ws = wb[sheet]
        ws.append(values)
        return ws.max_row

    def save(self, output: Optional[str] = None) -> Path:
        """Save the workbook"""
        return self._save(output)
